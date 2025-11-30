# brezin_app.py
import json
import os
import sys
from pathlib import Path
from datetime import datetime

import webview
from openpyxl import Workbook, load_workbook

# --- Paramètres de l'application ---
APP_NAME   = "Jeu du Brezin"        # Titre de la fenêtre pywebview
HTML_FILE  = "index.html"           # Fichier HTML à charger (UI du jeu)
EXCEL_NAME = "brezin_scores.xlsx"   # Classeur Excel où l'on stocke les résultats


# --- Fonctions utilitaires pour résoudre les chemins selon le contexte (dev vs packagé) ---
def get_assets_dir():
    """
    Renvoie le dossier où se trouvent les assets (index.html, JS/CSS).
    - Cas PyInstaller (onefile) : sys._MEIPASS (dossier temporaire de décompression)
    - Cas PyInstaller (onefolder) : dossier de l'exécutable
    - Cas développement : dossier du script courant
    """
    if hasattr(sys, "_MEIPASS"):
        return Path(sys._MEIPASS)
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def get_data_dir():
    """
    Dossier où stocker les données modifiables (scores, configs…).

    - En version packagée (exe) : %APPDATA%\\Jeu du Brezin
    - En dev : dossier du script (pour te simplifier la vie)
    """
    if getattr(sys, "frozen", False):
        # Cas exe PyInstaller : données dans le profil utilisateur
        base = Path(os.getenv("APPDATA")) / APP_NAME
    else:
        # Cas développement : à côté du script
        base = Path(__file__).resolve().parent

    base.mkdir(parents=True, exist_ok=True)
    return base


# --- Chemins calculés ---
ASSETS_DIR  = get_assets_dir()              # Dossier des assets (lecture)
HTML_PATH   = ASSETS_DIR / HTML_FILE        # Chemin complet vers index.html
EXCEL_PATH  = get_data_dir() / EXCEL_NAME   # Emplacement du classeur Excel


# --- Schéma de colonnes par défaut (évolutif) ---
HEADERS = [
    "date", "heure", "vainqueur", "ecart",
    "score_joueur", "score_ordi",
    "annonces_joueur", "annonces_ordi",
    "nb_as_joueur", "nb_dix_joueur", "points_10as_joueur",
    "nb_as_ordi", "nb_dix_ordi", "points_10as_ordi",
    "liste_annonces_joueur", "liste_annonces_ordi",
    "_submitted_at", "userName",
]


def _num_or_blank(v):
    """
    Normalise une valeur :
    - None -> "" (cellule vide)
    - nombres -> inchangés
    - strings numériques (y compris avec virgule) -> float/int
    - autre -> string trimée
    """
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if s == "":
        return ""
    try:
        f = float(s.replace(",", "."))
        return int(f) if f.is_integer() else f
    except Exception:
        return s


# --- API exposée à JavaScript via pywebview (window.pywebview.api) ---
class Api:
    def save_row(self, payload: str):
        """
        Reçoit un JSON (string) depuis JS, nettoie les valeurs, assure les colonnes,
        et ajoute une ligne au classeur Excel.
        """
        try:
            # 1) Parse + normalisation
            data = json.loads(payload) if payload else {}
            clean = {
                k: _num_or_blank(v.strip() if isinstance(v, str) else v)
                for k, v in data.items()
            }
            clean["_submitted_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # 2) Ouverture (ou création) du classeur + feuille active
            if EXCEL_PATH.exists():
                wb = load_workbook(EXCEL_PATH)
                ws = wb.active
                # Si le fichier est vide (rare), on pose les en-têtes
                if ws.max_row == 0:
                    ws.append(HEADERS)
            else:
                EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
                wb = Workbook()
                ws = wb.active
                ws.title = "scores"
                ws.append(HEADERS)

            # 3) Évolution du schéma : ajoute les colonnes manquantes (HEADERS + clés nouvelles)
            existing = [c.value for c in ws[1]]  # liste des titres d'en-tête existants
            changed = False
            # Colonnes de base
            for h in HEADERS:
                if h not in existing:
                    existing.append(h)
                    ws.cell(row=1, column=len(existing)).value = h
                    changed = True
            # Colonnes dynamiques rencontrées dans le payload
            for k in clean.keys():
                if k not in existing:
                    existing.append(k)
                    ws.cell(row=1, column=len(existing)).value = k
                    changed = True
            # Si modifié, on relit l'en-tête (sécurité)
            if changed:
                existing = [c.value for c in ws[1]]

            # 4) Construction et ajout de la ligne
            row = [clean.get(h, "") for h in existing]
            ws.append(row)
            wb.save(EXCEL_PATH)
            return {"ok": True, "message": f"Ligne ajoutée dans {EXCEL_PATH.name}"}

        except Exception as e:
            # Renvoi d'une erreur exploitable côté JS
            return {"ok": False, "message": f"Erreur: {e}"}

    def read_rows(self, limit: int = 300):
        """
        Lit les dernières 'limit' lignes et renvoie un tableau de dicts
        (clé = nom de colonne d'en-tête).
        """
        try:
            # Classeur absent -> vide
            if not EXCEL_PATH.exists():
                return {"ok": True, "rows": []}

            # Lecture en mode optimisé
            wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
            ws = wb.active
            # Pas de données -> vide
            if ws.max_row < 2:
                return {"ok": True, "rows": []}

            headers = [c.value for c in ws[1]]
            if not any(headers):
                return {"ok": True, "rows": []}

            # On ne lit que les N dernières lignes (performance)
            start_row = max(2, ws.max_row - limit + 1)
            rows = []
            for r in ws.iter_rows(
                min_row=start_row, max_row=ws.max_row, values_only=True
            ):
                row = {}
                for i, h in enumerate(headers):
                    if not h:
                        continue
                    row[h] = r[i] if i < len(r) else None
                rows.append(row)

            return {"ok": True, "rows": rows}
        except Exception as e:
            return {"ok": False, "message": f"Erreur lecture Excel: {e}"}


# --- Lancement de la fenêtre pywebview ---
def main():
    # Sécurité: s'assurer que l'HTML existe bien à l'endroit attendu
    if not HTML_PATH.exists():
        raise FileNotFoundError(f"Fichier HTML introuvable: {HTML_PATH}")

    api = Api()

    # Création d'une unique fenêtre, en FENÊTRE MAXIMISÉE (pas fullscreen)
    window = webview.create_window(
        APP_NAME,
        url=HTML_PATH.as_uri(),  # as_uri() pour schéma file:// absolu
        text_select=True,        # autoriser sélection de texte
        js_api=api,
        maximized=True,          # fenêtre ouverte en taille écran
        fullscreen=False,        # on force à ne PAS être en plein écran
    )

    def on_loaded():
        window.min_size = (1500, 700)

    window.events.loaded += on_loaded

    webview.start(debug=False)


if __name__ == "__main__":
    main()
